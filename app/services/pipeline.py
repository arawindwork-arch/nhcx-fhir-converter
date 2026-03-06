"""
Conversion Pipeline Orchestrator
==================================
Coordinates the end-to-end PDF → FHIR InsurancePlan Bundle conversion.

Pipeline stages:
  1. PDF Text Extraction (pdfplumber + OCR)
  2. LLM Structured Data Extraction (LangChain)
  3. FHIR R4 InsurancePlan Mapping (NRCeS v6.5.0)
  4. NRCeS Profile Validation
  5. Output Generation (JSON bundle + mapping Excel)
"""

import os
import json
import logging
from typing import Optional

from app.config import settings
from app.api.models import ConversionResponse, ValidationError as ValError
from app.services.pdf_extractor import PDFExtractor
from app.services.llm_processor import LLMProcessor
from app.services.fhir_mapper import FHIRMapper
from app.services.validator import NHCXValidator

logger = logging.getLogger("nhcx-converter.pipeline")


class ConversionPipeline:
    """Orchestrates the complete PDF → FHIR conversion pipeline."""

    def __init__(self):
        self.pdf_extractor = PDFExtractor()
        self.llm_processor = LLMProcessor()
        self.fhir_mapper = FHIRMapper()
        self.validator = NHCXValidator()

    async def run(
        self,
        job_id: str,
        pdf_path: str,
        source_filename: str,
        mapping_config: Optional[str] = None,
        insurer_name_override: Optional[str] = None,
        plan_type_override: Optional[str] = None,
        auto_validate: bool = True,
        skip_review: bool = False,
    ) -> ConversionResponse:
        """
        Execute the full conversion pipeline.

        Returns ConversionResponse with status:
          - "review"    → extracted data ready for human review
          - "completed" → FHIR bundle generated (if skip_review=True)
          - "failed"    → error occurred
        """
        logger.info("=" * 60)
        logger.info("Job %s: Starting conversion pipeline", job_id)
        logger.info("=" * 60)

        # ── Stage 1: PDF Text Extraction ─────────────────────────────────
        logger.info("Job %s: Stage 1 — Extracting text from PDF", job_id)
        try:
            extraction = self.pdf_extractor.extract(pdf_path)
            logger.info(
                "Job %s: Extracted %d pages, %d tables, quality=%s",
                job_id,
                extraction.total_pages,
                extraction.tables_found,
                extraction.extraction_quality,
            )
        except Exception as e:
            logger.error("Job %s: PDF extraction failed: %s", job_id, e)
            return ConversionResponse(
                job_id=job_id,
                status="failed",
                message=f"PDF extraction failed: {str(e)}",
            )

        # ── Stage 2: LLM Structured Extraction ──────────────────────────
        logger.info("Job %s: Stage 2 — LLM extraction (%s)", job_id, settings.LLM_MODEL)
        try:
            chunks = self.pdf_extractor.chunk_text(extraction.full_text)
            extracted_plan = await self.llm_processor.extract_plan_data(
                document_text=extraction.full_text,
                chunks=chunks,
                source_filename=source_filename,
            )

            # Apply overrides
            if insurer_name_override:
                extracted_plan.insurer_name = insurer_name_override
            if plan_type_override:
                extracted_plan.plan_type = plan_type_override

            logger.info(
                "Job %s: Extracted plan '%s' from '%s' "
                "(coverages=%d, exclusions=%d, costs=%d)",
                job_id,
                extracted_plan.plan_name,
                extracted_plan.insurer_name,
                len(extracted_plan.coverages),
                len(extracted_plan.exclusions),
                len(extracted_plan.plan_costs),
            )
        except Exception as e:
            logger.error("Job %s: LLM extraction failed: %s", job_id, e)
            return ConversionResponse(
                job_id=job_id,
                status="failed",
                message=f"LLM extraction failed: {str(e)}",
            )

        # ── Check if human review needed ─────────────────────────────────
        if not skip_review:
            # Save extracted data for review
            review_path = os.path.join(
                settings.REVIEW_DIR, f"{job_id}_extracted.json"
            )
            with open(review_path, "w") as f:
                json.dump(extracted_plan.model_dump(), f, indent=2, default=str)

            logger.info("Job %s: Queued for human review at %s", job_id, review_path)

            return ConversionResponse(
                job_id=job_id,
                status="review",
                message=(
                    "Extraction complete. Review extracted data at "
                    f"POST /api/v1/review/{job_id} before FHIR bundle generation."
                ),
                extracted_plan=extracted_plan,
            )

        # ── Stage 3: FHIR Mapping ───────────────────────────────────────
        logger.info("Job %s: Stage 3 — FHIR InsurancePlan mapping", job_id)
        try:
            fhir_bundle = self.fhir_mapper.map_to_insurance_plan_bundle(extracted_plan)
        except Exception as e:
            logger.error("Job %s: FHIR mapping failed: %s", job_id, e)
            return ConversionResponse(
                job_id=job_id,
                status="failed",
                message=f"FHIR mapping failed: {str(e)}",
                extracted_plan=extracted_plan,
            )

        # ── Stage 4: Validation ──────────────────────────────────────────
        validation_errors = []
        if auto_validate:
            logger.info("Job %s: Stage 4 — NRCeS profile validation", job_id)
            raw_errors = self.validator.validate(fhir_bundle)
            validation_errors = [ValError(**e) for e in raw_errors]

            error_count = sum(1 for e in validation_errors if e.severity == "error")
            if error_count > 0:
                logger.warning(
                    "Job %s: Validation found %d errors", job_id, error_count
                )

        # ── Stage 5: Output Generation ───────────────────────────────────
        logger.info("Job %s: Stage 5 — Generating output files", job_id)

        # Save FHIR bundle
        output_path = os.path.join(
            settings.OUTPUT_DIR, f"{job_id}_InsurancePlan_Bundle.json"
        )
        with open(output_path, "w") as f:
            json.dump(fhir_bundle, f, indent=2)

        # Generate mapping Excel
        try:
            self._generate_mapping_excel(job_id, extracted_plan, fhir_bundle)
        except Exception as e:
            logger.warning("Job %s: Mapping Excel generation failed: %s", job_id, e)

        logger.info("Job %s: Pipeline complete! Output: %s", job_id, output_path)

        return ConversionResponse(
            job_id=job_id,
            status="completed",
            message="FHIR InsurancePlan bundle generated successfully",
            extracted_plan=extracted_plan,
            fhir_bundle=fhir_bundle,
            validation_errors=validation_errors,
            output_file=output_path,
        )

    def _generate_mapping_excel(self, job_id, plan, bundle):
        """Generate FHIR Mapping Excel document for submission."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()

            # ── Sheet 1: Field Mapping ───────────────────────────────────
            ws = wb.active
            ws.title = "FHIR Mapping"

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill("solid", fgColor="1F4E79")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            headers = [
                "PDF Field / Section",
                "Extracted Value",
                "FHIR Resource",
                "FHIR Path",
                "FHIR Value / Code",
                "NRCeS Profile",
                "Notes",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # Populate mapping rows
            mappings = [
                ["Plan Name", plan.plan_name, "InsurancePlan", "InsurancePlan.name", plan.plan_name, "Required (1..1)", ""],
                ["Plan Identifier / UIN", plan.plan_identifier, "InsurancePlan", "InsurancePlan.identifier", plan.plan_identifier, "Required (1..1)", ""],
                ["Plan Type", plan.plan_type, "InsurancePlan", "InsurancePlan.type", plan.plan_type, "Required (1..1)", "ndhm-insuranceplan-type"],
                ["Status", plan.status, "InsurancePlan", "InsurancePlan.status", plan.status, "Required (1..1)", ""],
                ["Insurer Name", plan.insurer_name, "Organization", "Organization.name", plan.insurer_name, "Required (1..1)", "Referenced via ownedBy"],
                ["IRDAI Registration", plan.insurer_irdai_registration, "Organization", "Organization.identifier", plan.insurer_irdai_registration, "Optional", ""],
                ["TPA", plan.tpa_name, "Organization", "Organization.name (TPA)", plan.tpa_name, "Optional", "Referenced via administeredBy"],
                ["Effective From", plan.effective_from, "InsurancePlan", "InsurancePlan.period.start", plan.effective_from, "Required (1..1)", ""],
                ["Coverage Area", ", ".join(plan.coverage_areas), "InsurancePlan", "InsurancePlan.coverageArea", ", ".join(plan.coverage_areas), "Optional", ""],
            ]

            # Add coverage mappings
            for i, cov in enumerate(plan.coverages):
                mappings.append([
                    f"Coverage: {cov.coverage_type}",
                    cov.description or cov.coverage_type,
                    "InsurancePlan",
                    f"InsurancePlan.coverage[{i}].type",
                    cov.coverage_type,
                    "Required (1..*)",
                    "ndhm-coverage-type",
                ])
                for j, ben in enumerate(cov.benefits):
                    mappings.append([
                        f"  Benefit: {ben.name}",
                        ben.description or ben.name,
                        "InsurancePlan",
                        f"InsurancePlan.coverage[{i}].benefit[{j}]",
                        ben.name,
                        "Required (1..*)",
                        "ndhm-benefit-type",
                    ])

            # Add exclusion mappings
            for i, excl in enumerate(plan.exclusions):
                mappings.append([
                    f"Exclusion: {excl.name}",
                    excl.description or excl.name,
                    "InsurancePlan",
                    f"InsurancePlan.extension:Claim-Exclusion[{i}]",
                    excl.exclusion_type or "permanent",
                    "Optional (0..*)",
                    "Claim-Exclusion extension",
                ])

            for row_idx, row_data in enumerate(mappings, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                    cell.border = thin_border

            # Auto-fit column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

            # Save
            excel_path = os.path.join(
                settings.OUTPUT_DIR, f"{job_id}_FHIR_Mapping.xlsx"
            )
            wb.save(excel_path)
            logger.info("Mapping Excel saved: %s", excel_path)

        except ImportError:
            logger.warning("openpyxl not installed — skipping Excel generation")
